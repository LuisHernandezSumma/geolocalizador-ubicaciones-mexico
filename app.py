import streamlit as st
import pandas as pd
import folium
from streamlit_folium import st_folium
import plotly.express as px
import json, gzip, pickle, time, io, base64, html
from pathlib import Path
from utils.geocoder import geocode_row, enrich_zones, build_kml, geocode_single
from utils.data_loader import load_cp_lookup, load_kml_zones, load_zonas_display

# Paleta institucional SUMMA (extraida del tablero Power BI)
SUMMA_AZUL = "#003EA5"        # azul fuerte - headers, botones
SUMMA_AZUL_MEDIO = "#688BC6"  # azul medio - acentos
SUMMA_LAVANDA = "#ADBCDD"     # lavanda azul - barras secundarias
SUMMA_VERDE = "#C9E7DD"       # verde menta - seccion valores
SUMMA_ROSA = "#F2DBED"        # rosa claro - seccion alterna
SUMMA_GRIS = "#6B7280"        # texto gris
SUMMA_GRIS_CLARO = "#ECECEC"  # fondos
SUMMA_AZUL_CLARO = "#EaEFF7"  # fondo suave
SUMMA_GRIS_COMBO = "#D4D9DC"  # gris suave para combos/titulos
# Secuencia de colores para graficas multi-categoria
SUMMA_PALETA = ["#003EA5", "#688BC6", "#ADBCDD", "#C9E7DD", "#9BC4B5",
                "#F2DBED", "#D4A5C9", "#86A1CE", "#5B7BB4", "#B8C8E0"]

st.set_page_config(page_title="Geolocalizador SUMMA", page_icon=":world_map:", layout="wide")

st.markdown(f"""
<style>
.main-title {{ font-size: 1.9rem; font-weight: 700; color: {SUMMA_AZUL}; margin-bottom: 0; }}
.sub-title  {{ font-size: 0.95rem; color: {SUMMA_GRIS}; margin-bottom: 1.2rem; }}
/* Botones: fondo azul, TEXTO BLANCO forzado */
.stButton>button[kind="primary"], .stDownloadButton>button {{
    background-color: {SUMMA_AZUL} !important;
    border-color: {SUMMA_AZUL} !important;
    color: #FFFFFF !important;
    font-weight: 600 !important;
}}
.stButton>button[kind="primary"] *, .stDownloadButton>button * {{ color: #FFFFFF !important; }}
.stButton>button[kind="primary"]:hover, .stDownloadButton>button:hover {{
    background-color: {SUMMA_AZUL_MEDIO} !important;
    border-color: {SUMMA_AZUL_MEDIO} !important;
    color: #FFFFFF !important;
}}
.stProgress > div > div > div > div {{ background-color: {SUMMA_AZUL}; }}
div[data-baseweb="tab-list"] button[aria-selected="true"] {{ color: {SUMMA_AZUL}; font-weight: 600; }}
div[data-baseweb="tab-highlight"] {{ background-color: {SUMMA_AZUL}; }}
.summa-header {{ display:flex; align-items:center; gap:16px; margin-bottom:0.4rem;
                 border-bottom: 3px solid {SUMMA_AZUL}; padding-bottom: 10px; }}
[data-testid="stMetric"] {{ background: {SUMMA_AZUL_CLARO}; border-radius: 8px;
    padding: 10px 14px; border-left: 4px solid {SUMMA_AZUL}; }}
[data-testid="stMetricValue"] {{ color: {SUMMA_AZUL}; }}
/* Etiquetas de selectbox tipo chip (el emoji verde/blanco indica deteccion) */
.stSelectbox label {{
    background-color: {SUMMA_GRIS_COMBO};
    padding: 2px 10px; border-radius: 6px;
    font-weight: 600 !important; color: {SUMMA_AZUL} !important;
}}
/* Combos del mapeo: color segun deteccion via wrapper */
/* Por defecto (otros combos de la app) gris suave */
div[data-baseweb="select"] > div {{
    background-color: {SUMMA_GRIS_COMBO} !important;
    border-color: #B8BFC4 !important;
}}
/* Marcadores ocultos */
.mk-ok, .mk-no {{ display: none; }}
/* El contenedor que CONTIENE el marcador .mk-ok, su siguiente hermano (el combo) -> verde */
div[data-testid="stElementContainer"]:has(.mk-ok) + div[data-testid="stElementContainer"] div[data-baseweb="select"] > div {{
    background-color: #D7EFDD !important; border-color: #7FBF95 !important;
}}
div[data-testid="stElementContainer"]:has(.mk-no) + div[data-testid="stElementContainer"] div[data-baseweb="select"] > div {{
    background-color: #FFFFFF !important; border-color: #D0D0D0 !important;
}}
/* Etiquetas: negrita, el color de fondo se asigna por campo segun deteccion */
.stTextInput label, .stSlider label {{
    background-color: {SUMMA_GRIS_COMBO};
    padding: 2px 10px; border-radius: 6px;
    font-weight: 600 !important; color: {SUMMA_AZUL} !important;
}}
/* Campos de texto con fondo gris para que se distingan del fondo blanco */
.stTextInput input {{
    background-color: {SUMMA_GRIS_COMBO} !important;
    border: 1px solid #B8BFC4 !important;
    border-radius: 6px !important;
}}
.stTextInput input::placeholder {{ color: #7A7F85 !important; }}

</style>
""", unsafe_allow_html=True)

def get_logo_b64():
    p = Path(__file__).parent / "assets" / "summa_logo.png"
    if p.exists():
        return base64.b64encode(p.read_bytes()).decode()
    return None

logo = get_logo_b64()
logo_img = f'<img src="data:image/png;base64,{logo}" style="height:56px;background:white;border-radius:8px;padding:4px"/>' if logo else ''
st.markdown(f"""
<div style="background:{SUMMA_AZUL};border-radius:10px;padding:16px 24px;margin-bottom:18px;
            display:flex;align-items:center;gap:18px">
  {logo_img}
  <div>
    <div style="font-size:1.7rem;font-weight:700;color:#fff;line-height:1.1">Geolocalizador de Ubicaciones México</div>
    <div style="font-size:0.9rem;color:#C7D4E4">Coordenadas, zonas sísmicas, cresta e hidrometeorológicas &middot; Intermediario de Reaseguro</div>
  </div>
</div>
""", unsafe_allow_html=True)

@st.cache_resource(show_spinner="Cargando datos de referencia...")
def get_reference_data():
    return load_cp_lookup(), load_kml_zones(), load_zonas_display()

cp_lookup, kml_zones, zonas_display = get_reference_data()

with st.sidebar:
    if logo:
        st.markdown(f'<img src="data:image/png;base64,{logo}" style="width:160px;margin-bottom:12px"/>', unsafe_allow_html=True)
    st.markdown("### Configuración")
    st.markdown(f"**{len(cp_lookup):,}** CPs en base de referencia")

    # API key de Google leida de forma segura desde Secrets (nunca del codigo)
    try:
        GOOGLE_API_KEY = st.secrets.get("GOOGLE_MAPS_API_KEY", "")
    except Exception:
        GOOGLE_API_KEY = ""
    if GOOGLE_API_KEY:
        st.success("Google Maps: activo", icon="✅")
    else:
        st.info("Google Maps: no configurado (solo Nominatim)", icon="ℹ️")

    st.markdown("---")
    st.markdown("**Estrategia de búsqueda:**")
    st.markdown("1. Coordenadas existentes → inverso")
    st.markdown("2. CP + Estado")
    st.markdown("3. Solo CP")
    st.markdown("4. Nombre + Ciudad + Estado")
    st.markdown("5. Ciudad + Estado")
    st.markdown("6. Google Maps" + (" (activo)" if GOOGLE_API_KEY else " (inactivo)"))
    st.markdown("---")
    delay = st.slider("Delay entre requests (seg)", 1.0, 3.0, 1.1, 0.1,
                      help="Para respetar el rate limit de Nominatim")

tab_punto, tab_excel = st.tabs(["🔎 Consulta rápida (un punto)", "📊 Procesar Excel"])

with tab_punto:
    st.markdown("Escribe un **CP** o una **dirección** para conocer sus zonas (sísmica, cresta, huracán).")
    cpa, cpb = st.columns([4, 1])
    punto_txt = cpa.text_input("CP o dirección", key="punto_txt",
                               placeholder="Ej. 57510  ·  o  ·  Av. Reforma 222, CDMX",
                               label_visibility="collapsed")
    buscar = cpb.button("Buscar", type="primary", use_container_width=True)

    if buscar and punto_txt.strip():
        with st.spinner("Buscando..."):
            r = geocode_single(punto_txt, cp_lookup, kml_zones, delay, GOOGLE_API_KEY)
        if r.get('lat_geo'):
            st.success(f"Encontrado por: {r.get('metodo','')}")
            # Resultado resaltado en marco azul para que no pase desapercibido
            zs = r.get('zona_sismica') or "—"
            zc = r.get('zona_cresta') or "—"
            zh = r.get('hidro2') or "—"
            st.markdown(f"""
            <div style="border:2px solid {SUMMA_AZUL};border-radius:12px;
                        background:{SUMMA_AZUL_CLARO};padding:16px 18px;margin:8px 0 4px">
              <div style="display:flex;gap:14px;flex-wrap:wrap">
                <div style="flex:1;min-width:150px;background:#fff;border-radius:8px;
                            padding:10px 14px;border-left:5px solid {SUMMA_AZUL}">
                  <div style="font-size:12px;color:{SUMMA_GRIS};font-weight:600">ZONA SÍSMICA</div>
                  <div style="font-size:28px;font-weight:700;color:{SUMMA_AZUL}">{zs}</div>
                </div>
                <div style="flex:1;min-width:150px;background:#fff;border-radius:8px;
                            padding:10px 14px;border-left:5px solid {SUMMA_AZUL}">
                  <div style="font-size:12px;color:{SUMMA_GRIS};font-weight:600">ZONA CRESTA</div>
                  <div style="font-size:28px;font-weight:700;color:{SUMMA_AZUL}">{zc}</div>
                </div>
                <div style="flex:1;min-width:150px;background:#fff;border-radius:8px;
                            padding:10px 14px;border-left:5px solid {SUMMA_AZUL}">
                  <div style="font-size:12px;color:{SUMMA_GRIS};font-weight:600">ZONA HURACÁN</div>
                  <div style="font-size:28px;font-weight:700;color:{SUMMA_AZUL}">{zh}</div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)
            d1, d2, d3 = st.columns(3)
            d1.caption(f"**Estado:** {r.get('estado_geo') or '—'}")
            d2.caption(f"**Municipio:** {r.get('municipio_geo') or '—'}")
            d3.caption(f"**CP:** {r.get('cp_geo') or '—'}")
            try:
                latf, lngf = float(r['lat_geo']), float(r['lng_geo'])
                mp = folium.Map(location=[latf, lngf], zoom_start=14, tiles='CartoDB positron')
                folium.Marker([latf, lngf],
                    tooltip=punto_txt,
                    popup=folium.Popup(
                        f"<b>{punto_txt}</b><br>Sísmica: {r.get('zona_sismica','')}<br>"
                        f"Cresta: {r.get('zona_cresta','')}<br>Zona Huracán: {r.get('hidro2','')}",
                        max_width=240),
                    icon=folium.Icon(color='blue', icon='map-pin', prefix='fa')
                ).add_to(mp)
                st_folium(mp, width=None, height=400, returned_objects=[])
            except (ValueError, TypeError):
                st.info("Punto encontrado pero sin coordenadas para mapear.")
        else:
            st.warning(r.get('observacion', 'No se encontró el punto.'))
    elif buscar:
        st.warning("Escribe un CP o una dirección primero.")

with tab_excel:
    uploaded = st.file_uploader("Selecciona tu archivo Excel", type=["xlsx", "xls"],
        help="El archivo puede tener cualquier formato: dirección, CP, estado, ciudad, lat/lng")

    if uploaded:
        from openpyxl import load_workbook
        import unicodedata
        file_bytes = uploaded.read()
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        header_row = 0
        for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
            str_vals = [str(v).strip() for v in row
                        if v is not None and str(v).strip() not in ('', 'None', 'nan')
                        and not str(v).replace('.', '').replace('-', '').strip().isnumeric()]
            if len(str_vals) > 2:
                header_row = i
                break
        uploaded.seek(0)
        df = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
        df = df.dropna(axis=1, how='all')
        df.columns = [str(c).strip() for c in df.columns]
        st.success(f"Archivo cargado: **{uploaded.name}** - {len(df):,} filas, {len(df.columns)} columnas")

        with st.expander("Mapeo de columnas", expanded=True):
            cols = ["- no usar -"] + list(df.columns)
            CAMPOS = {
                'lat': ('📍 Latitud existente',  ['latitud', 'latitude', 'lat']),
                'lng': ('📍 Longitud existente', ['longitud', 'longitude', 'lon', 'lng']),
                'est': ('🗺️ Estado',       ['estado', 'state', 'entidad']),
                'mun': ('🏙️ Ciudad/Municipio', ['municipio', 'ciudad', 'city', 'poblacion', 'poblacion', 'ciudad juarez', 'locacion', 'locacion']),
                'cp':  ('📮 CP',                 ['cp', 'c.p.', 'c.p', 'codigo postal', 'codigo postal', 'postal', 'zip', 'cod postal']),
                'nom': ('🏢 Nombre/Sucursal/Edificio', ['nombre del puente', 'nombre del inmueble', 'nombre', 'sucursal', 'tienda', 'unidad']),
                'dir': ('📌 Dirección',          ['direccion', 'direccion', 'domicilio', 'address', 'ubicacion', 'ubicacion']),
                'neg': ('💼 Negocio',            ['negocio', 'grupo', 'asegurado', 'cliente', 'ramo', 'dependencia']),
                'mon': ('💱 Moneda',             ['moneda', 'divisa', 'currency']),
                'vinm':('🏛️ Valor Inmueble', ['valor inmueble', 'valor inm', 'inmueble', 'edificio', 'edificios', 'valor edificio', 'suma asegurada edificio']),
                'vcon':('📦 Valor Contenidos',   ['valor contenidos', 'valor con', 'contenidos', 'contenido', 'mobiliario', 'suma asegurada contenidos']),
                'vtot':('💰 Valor Total',        ['valor total', 'valor', 'tiv', 'tivs', 'suma asegurada', 'suma asegurada total', 'total asegurado']),
            }

            def _norm(c):
                nc = str(c).strip().lower()
                return ''.join(ch for ch in unicodedata.normalize('NFD', nc) if unicodedata.category(ch) != 'Mn')

            def autodetect(field_kw):
                for col in df.columns:
                    if _norm(col) in field_kw:
                        return col
                for kw in sorted(field_kw, key=len, reverse=True):
                    for col in df.columns:
                        if _norm(col).startswith(kw):
                            return col
                return None

            col1, col2 = st.columns(2)
            mapping = {}
            for i, (fid, (label, kws)) in enumerate(CAMPOS.items()):
                detected = autodetect(kws)
                default_idx = cols.index(detected) if detected and detected in cols else 0
                key = f"map_{fid}"
                # Estado actual del combo (refleja seleccion manual en vivo, no solo el auto-detect)
                current = st.session_state.get(key, detected if detected else "- no usar -")
                asignado = bool(current and current != "- no usar -")
                estado_ico = "\U0001F7E2" if asignado else "\u26AA"
                label_estado = f"{estado_ico} {label}"
                mk = "mk-ok" if asignado else "mk-no"
                with (col1 if i % 2 == 0 else col2):
                    # Marcador antes del combo; el CSS :has() colorea el combo hermano
                    st.markdown(f'<span class="{mk}"></span>', unsafe_allow_html=True)
                    sel = st.selectbox(label_estado, cols, index=default_idx, key=key)
                    mapping[fid] = sel if sel != "- no usar -" else None

        # Valores por defecto cuando Negocio/Moneda no vienen en columna
        cfg1, cfg2 = st.columns(2)
        negocio_default = cfg1.text_input("Negocio (si no viene en columna)", value="",
                                          placeholder="Ej. CONAGUA")
        moneda_default = cfg2.selectbox("Moneda por defecto", ["MXN", "USD", "EUR"], index=0)

        def gv(row, fid):
            col = mapping.get(fid)
            return str(row[col]).strip() if col and col in row and pd.notna(row[col]) else ''

        if len(df) > 0:
            row0 = df.iloc[0]
            lat0, lng0 = gv(row0, 'lat'), gv(row0, 'lng')
            has_coords = lat0 and lng0 and lat0 not in ('nan', '') and lng0 not in ('nan', '')
            cp0 = gv(row0, 'cp')
            with st.expander("Vista previa fila 1"):
                for fid, (label, _) in CAMPOS.items():
                    val = gv(row0, fid)
                    icon = "OK" if val and val != 'nan' else "-"
                    st.caption(f"{icon} {label}: **{val or '-'}**")
                if has_coords:
                    st.info(f"Modo: **inverso** → reverseGeocode({float(lat0):.5f}, {float(lng0):.5f})")
                elif cp0:
                    st.info(f"Modo: **CP en catálogo** → CP={cp0}")
                else:
                    nom = gv(row0, 'nom'); mun = gv(row0, 'mun'); est = gv(row0, 'est')
                    st.info(f"Modo: **Nominatim** → {', '.join(filter(None, [nom, mun, est, 'Mexico']))}")

        if st.button("Procesar", type="primary", use_container_width=True):
            results = []
            prog = st.progress(0, text="Iniciando...")
            log_area = st.empty()
            logs = []
            ok = warn = fail = 0
            for i, row in df.iterrows():
                prog.progress((i + 1) / len(df), text=f"Fila {i+1} de {len(df)}")
                result = geocode_row(row, mapping, cp_lookup, kml_zones, delay, GOOGLE_API_KEY)
                results.append(result)
                status = result.get('observacion', '')
                if 'CONFLICTO' in status: warn += 1
                elif 'No encontrado' in status or 'Sin datos' in status: fail += 1
                else: ok += 1
                logs.append(f"Fila {i+1}: {result.get('metodo','?')} -> {status[:60]}")
                log_area.code('\n'.join(logs[-8:]))
            prog.progress(1.0, text="Completado")
            st.session_state['results'] = results
            st.session_state['df_orig'] = df
            st.session_state['mapping'] = mapping
            st.session_state['negocio_default'] = negocio_default
            st.session_state['moneda_default'] = moneda_default
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total", len(df))
            c2.metric("OK", ok)
            c3.metric("Conflictos", warn)
            c4.metric("Sin resultado", fail)

    if 'results' in st.session_state:
        results = st.session_state['results']
        df_orig = st.session_state['df_orig']
        mapping = st.session_state['mapping']
        extra_cols = ['lat_geo', 'lng_geo', 'estado_geo', 'municipio_geo', 'cp_geo',
                      'zona_sismica', 'zona_cresta', 'hidro2', 'metodo', 'observacion']
        df_result = df_orig.copy()
        for col in extra_cols:
            df_result[col] = [r.get(col, '') for r in results]

        nom_col = mapping.get('nom')
        dir_col = mapping.get('dir')
        negocio_default = st.session_state.get('negocio_default', '')
        moneda_default = st.session_state.get('moneda_default', 'MXN')

        def to_num(v):
            """Convierte texto/numero a float; '' si no es numero."""
            if v is None:
                return None
            s_ = str(v).strip().replace('$', '').replace(',', '').replace(' ', '')
            if s_ in ('', 'nan', 'None'):
                return None
            try:
                return float(s_)
            except ValueError:
                return None

        def col_val(row, fid):
            c = mapping.get(fid)
            return row[c] if c and c in row and pd.notna(row[c]) else None

        # Construir columnas de valores
        v_inm, v_con, v_tot, v_neg, v_mon = [], [], [], [], []
        for _, row in df_orig.iterrows():
            inm = to_num(col_val(row, 'vinm'))
            con = to_num(col_val(row, 'vcon'))
            tot = to_num(col_val(row, 'vtot'))
            # Valor Total: si viene, se usa; si no, suma inm+con (lo que exista)
            if tot is None:
                partes = [x for x in (inm, con) if x is not None]
                tot = sum(partes) if partes else None
            v_inm.append(inm if inm is not None else '')
            v_con.append(con if con is not None else '')
            v_tot.append(tot if tot is not None else '')
            neg = col_val(row, 'neg')
            v_neg.append(str(neg).strip() if neg is not None else negocio_default)
            mon = col_val(row, 'mon')
            v_mon.append(str(mon).strip() if mon is not None else moneda_default)
        df_result['valor_inmueble'] = v_inm
        df_result['valor_contenidos'] = v_con
        df_result['valor_total'] = v_tot
        df_result['negocio'] = v_neg
        df_result['moneda'] = v_mon

        def row_search_blob(row):
            parts = [
                str(row[nom_col]) if nom_col and nom_col in row else '',
                str(row[dir_col]) if dir_col and dir_col in row else '',
                str(row.get('estado_geo', '')), str(row.get('municipio_geo', '')),
                str(row.get('cp_geo', '')),
            ]
            return ' '.join(parts).lower()

        df_result['_blob'] = df_result.apply(row_search_blob, axis=1)

        tab1, tab2, tab3, tab4 = st.tabs(["Mapa", "Tabla de resultados", "Dashboard zonas", "Dashboard valores"])

        with tab1:
            st.markdown("### Puntos geocodificados")
            search = st.text_input("Buscar por estado, CP, nombre/sucursal o dirección",
                                   placeholder="Escribe para filtrar...")

            def opts_with_count(col):
                d = df_result[df_result[col].astype(str).str.strip() != '']
                vc = d[col].value_counts()
                return [""] + [f"{k} ({v})" for k, v in vc.items()]

            def strip_count(label):
                return label.rsplit(" (", 1)[0] if label else ""

            fc1, fc2, fc3 = st.columns(3)
            fzs = fc1.selectbox("Zona Sísmica", opts_with_count('zona_sismica'), format_func=lambda x: x if x else "-")
            fzc = fc2.selectbox("Zona Cresta", opts_with_count('zona_cresta'), format_func=lambda x: x if x else "-")
            fh2 = fc3.selectbox("Zona Huracán", opts_with_count('hidro2'), format_func=lambda x: x if x else "-")

            df_map = df_result[df_result['lat_geo'] != ''].copy()
            df_map['lat_f'] = pd.to_numeric(df_map['lat_geo'], errors='coerce')
            df_map['lng_f'] = pd.to_numeric(df_map['lng_geo'], errors='coerce')
            df_map = df_map.dropna(subset=['lat_f', 'lng_f'])

            if strip_count(fzs): df_map = df_map[df_map['zona_sismica'] == strip_count(fzs)]
            if strip_count(fzc): df_map = df_map[df_map['zona_cresta'] == strip_count(fzc)]
            if strip_count(fh2): df_map = df_map[df_map['hidro2'] == strip_count(fh2)]
            if search.strip():
                terms = search.lower().split()
                df_map = df_map[df_map['_blob'].apply(lambda b: all(t in b for t in terms))]

            st.caption(f"Mostrando **{len(df_map)}** puntos")

            if len(df_map) > 0:
                m = folium.Map(location=[df_map['lat_f'].mean(), df_map['lng_f'].mean()],
                               zoom_start=5, tiles='CartoDB positron')

                # Capas de zonas (poligonos simplificados) con control on/off.
                # Solo "Zona Sísmica" encendida por defecto.
                _paleta = ['#003EA5', '#688BC6', '#ADBCDD', '#C9E7DD', '#9BC4B5',
                           '#F2DBED', '#D4A5C9', '#86A1CE', '#5B7BB4', '#E8A87C',
                           '#7FBF95', '#C45B5B']
                def _agregar_capa(cat_key, nombre_capa, encendida):
                    polys = zonas_display.get(cat_key, [])
                    if not polys:
                        return
                    fg = folium.FeatureGroup(name=nombre_capa, show=encendida)
                    nombres = sorted({p['name'] for p in polys})
                    color_de = {n: _paleta[i % len(_paleta)] for i, n in enumerate(nombres)}
                    for p in polys:
                        coords = p.get('polygon', [])
                        if len(coords) < 3:
                            continue
                        # coords vienen como [lng, lat]; Folium usa [lat, lng]
                        latlon = [[c[1], c[0]] for c in coords]
                        col = color_de.get(p['name'], '#003EA5')
                        folium.Polygon(locations=latlon, color=col, weight=1,
                                       fill=True, fill_color=col, fill_opacity=0.25,
                                       tooltip=f"{nombre_capa}: {p['name']}").add_to(fg)
                    fg.add_to(m)

                _agregar_capa('sismicas', 'Zona Sísmica', True)
                _agregar_capa('cresta', 'Zona Cresta', False)
                _agregar_capa('huracanes', 'Zona Huracán', False)

                # Marcadores de las ubicaciones (siempre visibles)
                fg_pts = folium.FeatureGroup(name='Ubicaciones', show=True)
                for _, row in df_map.iterrows():
                    obs = str(row.get('observacion', ''))
                    color = 'green' if obs.startswith('OK') else 'orange' if 'CONFLICTO' in obs else 'red'
                    nom = str(row[nom_col]) if nom_col and nom_col in row else ''
                    dir_ = str(row[dir_col]) if dir_col and dir_col in row else ''
                    popup_html = f"""
                    <div style="font-family:sans-serif;min-width:200px">
                      <b style="font-size:14px;color:{SUMMA_AZUL}">{html.escape(nom) or 'Sin nombre'}</b><br>
                      <span style="color:#666;font-size:12px">{html.escape(dir_)}</span><hr style="margin:6px 0">
                      <b>CP:</b> {row.get('cp_geo','')}<br>
                      <b>Estado:</b> {row.get('estado_geo','')}<br>
                      <b>Municipio:</b> {row.get('municipio_geo','')}<br>
                      <hr style="margin:6px 0">
                      <b>Zona Sísmica:</b> {row.get('zona_sismica','')}<br>
                      <b>Zona Cresta:</b> {row.get('zona_cresta','')}<br>
                      <b>Zona Huracán:</b> {row.get('hidro2','')}<br>
                      <hr style="margin:6px 0">
                      <span style="font-size:11px;color:#888">Método: {row.get('metodo','')}</span><br>
                      <span style="font-size:11px;color:#888">{html.escape(obs[:80])}</span>
                    </div>"""
                    folium.Marker(location=[row['lat_f'], row['lng_f']],
                                  popup=folium.Popup(popup_html, max_width=280),
                                  tooltip=nom or f"Fila {row.name+1}",
                                  icon=folium.Icon(color=color, icon='home', prefix='fa')).add_to(fg_pts)
                fg_pts.add_to(m)

                folium.LayerControl(collapsed=False).add_to(m)
                st_folium(m, width=None, height=550, returned_objects=[])
            else:
                st.warning("No hay puntos con coordenadas para mostrar con los filtros actuales.")

        with tab2:
            df_show = df_result.drop(columns=['_blob'])
            st.dataframe(df_show, use_container_width=True, height=400)
            cdl1, cdl2 = st.columns(2)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                df_show.to_excel(writer, index=False, sheet_name='Geocodificado')
            buf.seek(0)
            cdl1.download_button("Descargar Excel enriquecido", buf,
                file_name="resultado_geocodificado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary")
            kml_bytes = build_kml(df_result, nom_col, dir_col)
            cdl2.download_button("Descargar KML (Google My Maps)", kml_bytes,
                file_name="ubicaciones_summa.kml",
                mime="application/vnd.google-earth.kml+xml", use_container_width=True)
            st.caption("El KML se colorea por Zona Sísmica. Impórtalo en mymaps.google.com → Crear mapa → Importar.")

            st.markdown("---")
            st.markdown("**CSV para Power BI** (columnas genéricas y limpias)")
            # CSV generico: 14 columnas estandarizadas, una fila por punto
            nom_c = mapping.get('nom')
            csv_df = pd.DataFrame({
                'Nombre':   df_result[nom_c].astype(str) if nom_c and nom_c in df_result else '',
                'Negocio':  df_result['negocio'],
                'Estado':   df_result['estado_geo'],
                'Municipio':df_result['municipio_geo'],
                'CP':       df_result['cp_geo'].astype(str),
                'Lat':      df_result['lat_geo'],
                'Lon':      df_result['lng_geo'],
                'Valor_Inmueble':   df_result['valor_inmueble'],
                'Valor_Contenidos': df_result['valor_contenidos'],
                'Valor_Total':      df_result['valor_total'],
                'Moneda':   df_result['moneda'],
                'Zona_Sismica': df_result['zona_sismica'],
                'Zona_Cresta':  df_result['zona_cresta'],
                'Hidro2':       df_result['hidro2'],
            })
            csv_bytes = csv_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
            st.download_button("Descargar CSV para Power BI", csv_bytes,
                file_name="datos_powerbi.csv", mime="text/csv", use_container_width=True)
            st.caption("Sube este CSV a la carpeta que lee Power BI. Codificación UTF-8 con BOM para acentos.")

        with tab3:
            st.markdown("### Resumen por zonas y estados")
            total = len(df_result)
            con_coords = (df_result['lat_geo'] != '').sum()
            conflictos = df_result['observacion'].astype(str).str.contains('CONFLICTO').sum()
            sin_datos = df_result['observacion'].astype(str).str.contains('Sin datos').sum()
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total ubicaciones", total)
            m2.metric("Con coordenadas", int(con_coords))
            m3.metric("Conflictos", int(conflictos))
            m4.metric("Sin datos", int(sin_datos))
            st.markdown("---")

            def bar(col, titulo):
                d = df_result[df_result[col].astype(str).str.strip() != '']
                if len(d) == 0:
                    st.info(f"Sin datos para {titulo}")
                    return
                vc = d[col].value_counts().reset_index()
                vc.columns = [titulo, 'Ubicaciones']
                fig = px.bar(vc, x=titulo, y='Ubicaciones', text='Ubicaciones',
                             color=titulo, color_discrete_sequence=SUMMA_PALETA)
                fig.update_traces(textposition='outside')
                fig.update_layout(height=320, margin=dict(t=30, b=10, l=10, r=10),
                                  plot_bgcolor='white', showlegend=False,
                                  font=dict(color="#333"))
                st.plotly_chart(fig, use_container_width=True)

            g1, g2 = st.columns(2)
            with g1:
                st.markdown("**Por Zona Sísmica**"); bar('zona_sismica', 'Zona Sísmica')
            with g2:
                st.markdown("**Por Zona Cresta**"); bar('zona_cresta', 'Zona Cresta')
            g3, g4 = st.columns(2)
            with g3:
                st.markdown("**Por Zona Huracán**"); bar('hidro2', 'Zona Huracán')
            with g4:
                st.markdown("**Por Estado**"); bar('estado_geo', 'Estado')

        # ════════════════════════════ TAB 4: DASHBOARD VALORES ═══════════════════
        with tab4:
            st.markdown("### Valor Total asegurado por zonas y estados")

            dv = df_result.copy()
            dv['vt'] = pd.to_numeric(dv['valor_total'], errors='coerce')
            dv = dv.dropna(subset=['vt'])

            if len(dv) == 0:
                st.info("No hay valores numericos para graficar. Mapea la columna Valor Total (o Inmueble/Contenidos) al procesar.")
            else:
                moneda_lbl = dv['moneda'].mode().iloc[0] if len(dv['moneda'].mode()) else 'MXN'
                total_val = dv['vt'].sum()
                m1, m2 = st.columns(2)
                m1.metric("Valor Total asegurado", f"${total_val:,.0f} {moneda_lbl}")
                m2.metric("Ubicaciones con valor", len(dv))
                st.markdown("---")

                def bar_valor(col, titulo, horizontal=False):
                    d = dv[dv[col].astype(str).str.strip() != '']
                    if len(d) == 0:
                        st.info(f"Sin datos para {titulo}")
                        return
                    g = d.groupby(col)['vt'].sum().reset_index().sort_values('vt', ascending=False)
                    g.columns = [titulo, 'Valor']
                    if horizontal:
                        fig = px.bar(g, y=titulo, x='Valor', orientation='h',
                                     color=titulo, color_discrete_sequence=SUMMA_PALETA)
                        fig.update_layout(yaxis={'categoryorder': 'total ascending'})
                    else:
                        fig = px.bar(g, x=titulo, y='Valor',
                                     color=titulo, color_discrete_sequence=SUMMA_PALETA)
                    fig.update_layout(height=340, margin=dict(t=30, b=10, l=10, r=10),
                                      plot_bgcolor='white', showlegend=False, font=dict(color="#333"))
                    st.plotly_chart(fig, use_container_width=True)

                st.markdown("**Valor Total por Estado**")
                bar_valor('estado_geo', 'Estado', horizontal=True)

                v1, v2, v3 = st.columns(3)
                with v1:
                    st.markdown("**Por Zona Sísmica**"); bar_valor('zona_sismica', 'Zona Sísmica')
                with v2:
                    st.markdown("**Por Zona Cresta**"); bar_valor('zona_cresta', 'Zona Cresta')
                with v3:
                    st.markdown("**Por Zona Huracán**"); bar_valor('hidro2', 'Zona Huracán')